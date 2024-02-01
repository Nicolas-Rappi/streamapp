from io import BytesIO
from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, NamedStyle, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from streamlit import cache_data, secrets
from zipfile import ZipFile, ZipInfo


class ReportGenerator:
    date = date.today().isoformat()
    # stylish 
    totals_style = NamedStyle(name="totals_style")
    totals_style.font = Font(name='Arial', size=13, bold=True)
    totals_style.border = Border(top=Side(border_style='thick', color='FF000000'),bottom=Side(border_style=None, color='FF000000'))
    col = get_column_letter

    @classmethod
    @cache_data(ttl=timedelta(minutes=30))
    def generate_from_template(_cls, file_name:str, dfs:dict, sub_folder:str = '', base_file:str ='_', xslx_style:str = 'TableStyleMedium3'):
        """
        Return an .xlsx file
        dfs is a dict of names and Dataframes per sheets ex: {'Hoja 1': Dataframe1, 'ingresos': ingresos_dataframe}
        kargs: variables to define a summary

        """
        try:
            workbook = load_workbook(Path(secrets.utils_file, sub_folder, base_file +'.xlsx').as_posix())
        except:
            workbook = Workbook()
            for n, i in enumerate(dfs.keys()):
                if n == 0:
                    workbook.active.title = i
                else:
                    workbook.create_sheet(title=i)

        for name, df in dfs.items():
            # seguro de fecha para generar excel    
            for i in df.columns: 
                try: df[i] = df[i].dt.date 
                except: pass
            (max_row, max_col) = df.shape
            worksheet = workbook[name]
                
            for i in range(1,max_col+1):
                worksheet.column_dimensions[_cls.col(i)].width = 23
            
            tab = Table(displayName=name, ref=f"A1:{_cls.col(max_col)}{max_row+1}")
            tab.tableStyleInfo = TableStyleInfo(name=xslx_style)
            worksheet._tables.add(tab)
            
            header = base_file == '_'
            for c, r in enumerate(dataframe_to_rows(df, index=False, header=header),int(not header)): 
                for i, j in enumerate(r, 1):
                    worksheet.cell(c+1,i,j)

        with BytesIO() as buffer:
            workbook.save(buffer)
            file = buffer.getvalue()
        workbook.close()

        return [file_name+'.xlsx', file]
 

class InMemoryZip(object):
    zip_file = BytesIO()

    @classmethod
    def create_zip(cls, reports:list):
        """
        returns: zip archive
        """ 
        with ZipFile(cls.zip_file, 'w') as  zip_archive:
            for i in reports:
                zip_archive.writestr(
                    zinfo_or_arcname=ZipInfo(i[0]),
                    data=i[1]
                )
        
        return cls.zip_file